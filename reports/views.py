import io
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Count, Q, Avg, Sum
from django.utils import timezone
from datetime import datetime, timedelta
from io import BytesIO
from django.contrib import messages
from data_entry.models import DataEntry
from indicators.models import Indicator, ThematicArea
from core.models import County, Quarter
from users.decorators import view_reports_required, admin_required, ncpd_or_admin_required

# @login_required
# @view_reports_required
# def dashboard(request):
#     """Main dashboard with RBAC and user-specific data"""
#     user = request.user
    
#     # ===== RBAC: DETERMINE USER'S DATA SCOPE =====
#     if user.role and user.role.name == 'county_me' and user.county:
#         # County users only see their county
#         county = user.county
#         entries = DataEntry.objects.filter(county=county, status='approved')
#         all_entries = DataEntry.objects.filter(county=county)  # All statuses for this county
#         counties = County.objects.filter(id=county.id)
#         user_scope = f"County: {county.name}"
#         is_county_user = True
#         county_name = county.name
#     elif user.is_superuser or (user.role and user.role.name in ['admin', 'ncpd_me', 'policy_maker']):
#         # Admin/NCPD/Policy Maker see all data
#         entries = DataEntry.objects.filter(status='approved')
#         all_entries = DataEntry.objects.all()  # All statuses
#         counties = County.objects.filter(is_active=True)
#         user_scope = "All Counties"
#         is_county_user = False
#         county_name = None
#         county = None
#     else:
#         # Fallback
#         entries = DataEntry.objects.none()
#         all_entries = DataEntry.objects.none()
#         counties = County.objects.none()
#         user_scope = "No Access"
#         is_county_user = False
#         county_name = None
#         county = None
    
#     # ===== STATISTICS - ALL COUNTY-SPECIFIC =====
#     total_indicators = Indicator.objects.filter(is_active=True).count()
    
#     # For county users, only count indicators that have data for their county
#     if is_county_user and county:
#         county_indicator_ids = entries.values_list('indicator_id', flat=True).distinct()
#         total_indicators_with_data = Indicator.objects.filter(
#             is_active=True, 
#             id__in=county_indicator_ids
#         ).count()
#         total_indicators = total_indicators_with_data if total_indicators_with_data > 0 else Indicator.objects.filter(is_active=True).count()
    
#     total_counties = counties.count()
#     total_entries = entries.count()  # Only approved for this scope
#     approved_entries = entries.filter(status='approved').count()
    
#     # Pending approvals - for county users, only their county's pending
#     if is_county_user and county:
#         pending_approvals = DataEntry.objects.filter(county=county, status='submitted').count()
#     else:
#         pending_approvals = DataEntry.objects.filter(status='submitted').count()
    
#     # ===== SUBMISSION RATE - County specific =====
#     current_quarter = Quarter.objects.filter(is_active=True, is_closed=False).first()
#     submission_rate = 0
#     total_counties_for_quarter = 0
#     submitted_counties = 0
    
#     if current_quarter:
#         if is_county_user and county:
#             # For county users: check if they've submitted for current quarter
#             has_submitted = DataEntry.objects.filter(
#                 county=county,
#                 quarter=current_quarter,
#                 status__in=['submitted', 'approved']
#             ).exists()
#             submission_rate = 100 if has_submitted else 0
#         else:
#             # For admin: all counties
#             total_counties_for_quarter = County.objects.filter(is_active=True).count()
#             submitted_counties = DataEntry.objects.filter(
#                 quarter=current_quarter,
#                 status__in=['submitted', 'approved']
#             ).values('county').distinct().count()
#             submission_rate = round((submitted_counties / total_counties_for_quarter * 100) if total_counties_for_quarter > 0 else 0)
    
#     # ===== OVERALL PERFORMANCE =====
#     if total_entries > 0:
#         # Calculate how many entries met their target
#         met_target = 0
#         for entry in entries:
#             if entry.is_met() is True:
#                 met_target += 1
#         overall_performance = round((met_target / total_entries * 100))
#     else:
#         overall_performance = 0
    
#     # ===== COUNTIES WITH NO DATA =====
#     if is_county_user and county:
#         # For county users: check if they have any data
#         has_data = entries.exists()
#         counties_with_no_data = 0 if has_data else 1
#     else:
#         # For admin: count counties with no approved data
#         all_counties = County.objects.filter(is_active=True)
#         counties_with_data = DataEntry.objects.filter(
#             status='approved'
#         ).values('county').distinct()
#         counties_with_data_ids = [c['county'] for c in counties_with_data]
#         counties_with_no_data = all_counties.exclude(id__in=counties_with_data_ids).count()
    
#     # ===== THEMATIC PERFORMANCE =====
#     thematic_performance = []
#     thematic_colors = {
#         'Fertility': '#1a5632',
#         'Morbidity & Mortality': '#b71c1c',
#         'Migration & Urbanization': '#f39c12',
#         'PHED': '#2d8a4e'
#     }
    
#     for area in ThematicArea.objects.all():
#         area_indicators = Indicator.objects.filter(thematic_area=area, is_active=True)
#         area_entries = entries.filter(indicator__in=area_indicators)
#         total = area_entries.count()
#         met = 0
#         not_met = 0
#         for e in area_entries:
#             if e.is_met() is True:
#                 met += 1
#             elif e.is_met() is False:
#                 not_met += 1
        
#         # For county users, only show thematic areas with data
#         if is_county_user and total == 0:
#             continue
        
#         thematic_performance.append({
#             'name': area.name,
#             'code': area.code,
#             'total_indicators': area_indicators.count(),
#             'total_entries': total,
#             'met': met,
#             'not_met': not_met if total > 0 else 0,
#             'no_data': area_indicators.count() - total if area_indicators.count() > total else 0,
#             'percentage': round((met / total * 100) if total > 0 else 0),
#             'color': thematic_colors.get(area.name, '#6c757d')
#         })
    
#     # ===== CHART DATA =====
#     quarters = Quarter.objects.filter(is_active=True).order_by('-start_date')[:6]
#     chart_labels = []
#     chart_data = []
#     chart_target = []
    
#     for q in reversed(quarters):
#         chart_labels.append(q.name)
#         q_entries = entries.filter(quarter=q)
#         q_total = q_entries.count()
#         q_met = 0
#         for e in q_entries:
#             if e.is_met() is True:
#                 q_met += 1
#         chart_data.append(round((q_met / q_total * 100) if q_total > 0 else 0))
#         chart_target.append(65)
    
#     if not chart_labels:
#         chart_labels = ['No Data']
#         chart_data = [0]
#         chart_target = [65]
    
#     # ===== STATUS DISTRIBUTION =====
#     if is_county_user and county:
#         # County users only see their county's status distribution
#         status_data = {
#             'approved': DataEntry.objects.filter(county=county, status='approved').count(),
#             'submitted': DataEntry.objects.filter(county=county, status='submitted').count(),
#             'rejected': DataEntry.objects.filter(county=county, status='rejected').count(),
#             'draft': DataEntry.objects.filter(county=county, status='draft').count(),
#         }
#     else:
#         status_data = {
#             'approved': DataEntry.objects.filter(status='approved').count(),
#             'submitted': DataEntry.objects.filter(status='submitted').count(),
#             'rejected': DataEntry.objects.filter(status='rejected').count(),
#             'draft': DataEntry.objects.filter(status='draft').count(),
#         }
    
#     # ===== COUNTY PERFORMANCE =====
#     county_performance = []
#     for c in counties:
#         county_entries = entries.filter(county=c)
#         total = county_entries.count()
#         if total > 0:
#             met = 0
#             for e in county_entries:
#                 if e.is_met() is True:
#                     met += 1
#             county_performance.append({
#                 'name': c.name,
#                 'total': total,
#                 'met': met,
#                 'percentage': round((met / total * 100) if total > 0 else 0)
#             })
    
#     county_performance.sort(key=lambda x: x['percentage'], reverse=True)
#     county_performance = county_performance[:10]
    
#     # ===== RECENT ACTIVITY =====
#     recent_entries = entries.select_related('county', 'quarter', 'indicator', 'indicator__thematic_area').order_by('-created_at')[:10]
    
#     context = {
#         # Stats
#         'total_indicators': total_indicators,
#         'total_counties': total_counties,
#         'total_entries': total_entries,
#         'approved_entries': approved_entries,
#         'pending_approvals': pending_approvals,
#         'submission_rate': submission_rate,
#         'current_quarter': current_quarter,
#         'overall_performance': overall_performance,
#         'counties_with_no_data': counties_with_no_data,
#         'user_scope': user_scope,
        
#         # Thematic performance
#         'thematic_performance': thematic_performance,
        
#         # Chart data (as JSON)
#         'chart_labels': json.dumps(chart_labels),
#         'chart_data': json.dumps(chart_data),
#         'chart_target': json.dumps(chart_target),
        
#         # Status data
#         'status_data': status_data,
        
#         # County performance
#         'county_performance': county_performance,
        
#         # Recent entries
#         'recent_entries': recent_entries,
        
#         # User info
#         'user_role': user.role.name if user.role else 'No Role',
#         'is_superuser': user.is_superuser,
#         'is_county_user': is_county_user,
#         'county_name': county_name,
#     }
    
#     return render(request, 'reports/dashboard.html', context)

@login_required
@view_reports_required
def dashboard(request):
    """Main dashboard with RBAC and user-specific data"""
    user = request.user
    
    # ===== RBAC: DETERMINE USER'S DATA SCOPE =====
    if user.role and user.role.name == 'county_me' and user.county:
        county = user.county
        entries = DataEntry.objects.filter(county=county, status='approved')
        all_entries = DataEntry.objects.filter(county=county)
        counties = County.objects.filter(id=county.id)
        user_scope = f"County: {county.name}"
        is_county_user = True
        county_name = county.name
    elif user.is_superuser or (user.role and user.role.name in ['admin', 'ncpd_me', 'policy_maker']):
        entries = DataEntry.objects.filter(status='approved')
        all_entries = DataEntry.objects.all()
        counties = County.objects.filter(is_active=True)
        user_scope = "All Counties"
        is_county_user = False
        county_name = None
        county = None
    else:
        entries = DataEntry.objects.none()
        all_entries = DataEntry.objects.none()
        counties = County.objects.none()
        user_scope = "No Access"
        is_county_user = False
        county_name = None
        county = None
    
    # ===== STATISTICS =====
    total_indicators = Indicator.objects.filter(is_active=True).count()
    
    if is_county_user and county:
        county_indicator_ids = entries.values_list('indicator_id', flat=True).distinct()
        total_indicators_with_data = Indicator.objects.filter(
            is_active=True, 
            id__in=county_indicator_ids
        ).count()
        total_indicators = total_indicators_with_data if total_indicators_with_data > 0 else Indicator.objects.filter(is_active=True).count()
    
    total_counties = counties.count()
    total_entries = entries.count()
    approved_entries = entries.filter(status='approved').count()
    
    if is_county_user and county:
        pending_approvals = DataEntry.objects.filter(county=county, status='submitted').count()
    else:
        pending_approvals = DataEntry.objects.filter(status='submitted').count()
    
    # ===== SUBMISSION RATE =====
    current_quarter = Quarter.objects.filter(is_active=True, is_closed=False).first()
    submission_rate = 0
    if current_quarter:
        if is_county_user and county:
            has_submitted = DataEntry.objects.filter(
                county=county,
                quarter=current_quarter,
                status__in=['submitted', 'approved']
            ).exists()
            submission_rate = 100 if has_submitted else 0
        else:
            total_counties_for_quarter = County.objects.filter(is_active=True).count()
            submitted_counties = DataEntry.objects.filter(
                quarter=current_quarter,
                status__in=['submitted', 'approved']
            ).values('county').distinct().count()
            submission_rate = round((submitted_counties / total_counties_for_quarter * 100) if total_counties_for_quarter > 0 else 0)
    
    # ===== OVERALL PERFORMANCE =====
    if total_entries > 0:
        met_target = 0
        for entry in entries:
            if entry.is_met() is True:
                met_target += 1
        overall_performance = round((met_target / total_entries * 100))
    else:
        overall_performance = 0
    
    # ===== COUNTIES WITH NO DATA =====
    if is_county_user and county:
        has_data = entries.exists()
        counties_with_no_data = 0 if has_data else 1
    else:
        all_counties = County.objects.filter(is_active=True)
        counties_with_data = DataEntry.objects.filter(
            status='approved'
        ).values('county').distinct()
        counties_with_data_ids = [c['county'] for c in counties_with_data]
        counties_with_no_data = all_counties.exclude(id__in=counties_with_data_ids).count()
    
    # ===== THEMATIC PERFORMANCE =====
    thematic_performance = []
    thematic_colors = {
        'Fertility': '#1a5632',
        'Morbidity & Mortality': '#b71c1c',
        'Migration & Urbanization': '#f39c12',
        'PHED': '#2d8a4e'
    }
    
    for area in ThematicArea.objects.all():
        area_indicators = Indicator.objects.filter(thematic_area=area, is_active=True)
        area_entries = entries.filter(indicator__in=area_indicators)
        total = area_entries.count()
        met = 0
        not_met = 0
        for e in area_entries:
            if e.is_met() is True:
                met += 1
            elif e.is_met() is False:
                not_met += 1
        
        if is_county_user and total == 0:
            continue
        
        thematic_performance.append({
            'name': area.name,
            'code': area.code,
            'total_indicators': area_indicators.count(),
            'total_entries': total,
            'met': met,
            'not_met': not_met if total > 0 else 0,
            'no_data': area_indicators.count() - total if area_indicators.count() > total else 0,
            'percentage': round((met / total * 100) if total > 0 else 0),
            'color': thematic_colors.get(area.name, '#6c757d')
        })
    
    # ===== QUARTERLY PERFORMANCE DATA FOR CHART =====
    quarters = Quarter.objects.filter(is_active=True).order_by('-start_date')[:8]
    chart_labels = []
    chart_data = []
    chart_target = []
    chart_quarters_data = []
    
    for q in reversed(quarters):
        chart_labels.append(q.name)
        q_entries = entries.filter(quarter=q)
        q_total = q_entries.count()
        q_met = 0
        for e in q_entries:
            if e.is_met() is True:
                q_met += 1
        percentage = round((q_met / q_total * 100) if q_total > 0 else 0)
        chart_data.append(percentage)
        chart_target.append(65)
        chart_quarters_data.append({
            'name': q.name,
            'total': q_total,
            'met': q_met,
            'percentage': percentage
        })
    
    if not chart_labels:
        chart_labels = ['No Data']
        chart_data = [0]
        chart_target = [65]
    
    # ===== STATUS DISTRIBUTION =====
    if is_county_user and county:
        status_data = {
            'approved': DataEntry.objects.filter(county=county, status='approved').count(),
            'submitted': DataEntry.objects.filter(county=county, status='submitted').count(),
            'rejected': DataEntry.objects.filter(county=county, status='rejected').count(),
            'draft': DataEntry.objects.filter(county=county, status='draft').count(),
        }
    else:
        status_data = {
            'approved': DataEntry.objects.filter(status='approved').count(),
            'submitted': DataEntry.objects.filter(status='submitted').count(),
            'rejected': DataEntry.objects.filter(status='rejected').count(),
            'draft': DataEntry.objects.filter(status='draft').count(),
        }
    
    # ===== COUNTY PERFORMANCE =====
    county_performance = []
    for c in counties:
        county_entries = entries.filter(county=c)
        total = county_entries.count()
        if total > 0:
            met = 0
            for e in county_entries:
                if e.is_met() is True:
                    met += 1
            county_performance.append({
                'name': c.name,
                'total': total,
                'met': met,
                'percentage': round((met / total * 100) if total > 0 else 0)
            })
    
    county_performance.sort(key=lambda x: x['percentage'], reverse=True)
    county_performance = county_performance[:10]
    
    # ===== RECENT ACTIVITY =====
    recent_entries = entries.select_related('county', 'quarter', 'indicator', 'indicator__thematic_area').order_by('-created_at')[:10]
    
    # ===== ADDITIONAL ANALYTICS =====
    # Most active counties (with most submissions)
    most_active_counties = DataEntry.objects.filter(status='approved').values('county__name').annotate(
        total=Count('id')
    ).order_by('-total')[:5]
    
    # Indicator completion rate
    indicator_completion = {}
    for ind in Indicator.objects.filter(is_active=True)[:10]:
        ind_entries = entries.filter(indicator=ind)
        if ind_entries.count() > 0:
            indicator_completion[ind.code] = {
                'name': ind.name,
                'total': ind_entries.count(),
                'met': sum(1 for e in ind_entries if e.is_met() is True),
                'percentage': round((sum(1 for e in ind_entries if e.is_met() is True) / ind_entries.count() * 100) if ind_entries.count() > 0 else 0)
            }
    
    context = {
        # Stats
        'total_indicators': total_indicators,
        'total_counties': total_counties,
        'total_entries': total_entries,
        'approved_entries': approved_entries,
        'pending_approvals': pending_approvals,
        'submission_rate': submission_rate,
        'current_quarter': current_quarter,
        'overall_performance': overall_performance,
        'counties_with_no_data': counties_with_no_data,
        'user_scope': user_scope,
        
        # Thematic performance
        'thematic_performance': thematic_performance,
        
        # Chart data (as JSON)
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
        'chart_target': json.dumps(chart_target),
        'chart_quarters_data': chart_quarters_data,
        
        # Status data
        'status_data': status_data,
        
        # County performance
        'county_performance': county_performance,
        
        # Recent entries
        'recent_entries': recent_entries,
        
        # Additional analytics
        'most_active_counties': most_active_counties,
        'indicator_completion': indicator_completion,
        
        # User info
        'user_role': user.role.name if user.role else 'No Role',
        'is_superuser': user.is_superuser,
        'is_county_user': is_county_user,
        'county_name': county_name,
    }
    
    return render(request, 'reports/dashboard.html', context)

@login_required
@view_reports_required
def report_list(request):
    """List all available reports - RBAC enforced"""
    user = request.user
    
    # Base querysets
    thematic_areas = ThematicArea.objects.all()
    indicators = Indicator.objects.filter(is_active=True)
    quarters = Quarter.objects.filter(is_active=True)
    
    # ===== RBAC: COUNTY USERS ONLY SEE THEIR COUNTY =====
    if user.role and user.role.name == 'county_me' and user.county:
        # County users only see their county
        counties = County.objects.filter(id=user.county.id)
        # Get data entries for their county only
        county_entries = DataEntry.objects.filter(county=user.county, status='approved')
        # Get indicators that have data for this county
        indicators = indicators.filter(
            id__in=county_entries.values_list('indicator_id', flat=True).distinct()
        )
        # Get thematic areas that have indicators with data
        thematic_areas = thematic_areas.filter(
            id__in=indicators.values_list('thematic_area_id', flat=True).distinct()
        )
        # Get quarters that have data for this county
        quarters = quarters.filter(
            id__in=county_entries.values_list('quarter_id', flat=True).distinct()
        )
    else:
        # Admin/NCPD/Policy Maker see all counties
        counties = County.objects.filter(is_active=True)
    
    # Get pending approvals count for admin/ncpd (only if user has permission)
    pending_count = 0
    if user.is_superuser or (user.role and user.role.name in ['admin', 'ncpd_me']):
        pending_count = DataEntry.objects.filter(status='submitted').count()
    
    context = {
        'thematic_areas': thematic_areas,
        'counties': counties,
        'quarters': quarters,
        'indicators': indicators,
        'pending_count': pending_count,
        'user_role': user.role.name if user.role else 'No Role',
        'is_county_user': user.role and user.role.name == 'county_me',
        'county_name': user.county.name if user.county else None,
    }
    return render(request, 'reports/list.html', context)

@login_required
@view_reports_required
def generate_report(request):
    """Generate custom report based on filters - APPROVED ONLY"""
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        county_id = request.POST.get('county')
        quarter_id = request.POST.get('quarter')
        thematic_area_id = request.POST.get('thematic_area')
        indicator_id = request.POST.get('indicator')
        format_type = request.POST.get('format', 'html')
        
        # Build query - ONLY APPROVED
        query = Q(status='approved')
        if county_id:
            query &= Q(county_id=county_id)
        if quarter_id:
            query &= Q(quarter_id=quarter_id)
        if thematic_area_id:
            indicators = Indicator.objects.filter(thematic_area_id=thematic_area_id)
            query &= Q(indicator__in=indicators)
        if indicator_id:
            query &= Q(indicator_id=indicator_id)
        
        # RBAC: County users only see their county
        if request.user.role == 'county_me' and request.user.county:
            query &= Q(county=request.user.county)
        
        entries = DataEntry.objects.filter(query).order_by('-created_at')
        
        if format_type == 'csv':
            return export_csv(entries)
        elif format_type == 'json':
            return export_json(entries)
        else:
            return render(request, 'reports/result.html', {'entries': entries})
    
    # GET - show form
    counties = County.objects.filter(is_active=True)
    quarters = Quarter.objects.filter(is_active=True)
    thematic_areas = ThematicArea.objects.all()
    indicators = Indicator.objects.filter(is_active=True)
    
    # RBAC: County users only see their county
    if request.user.role == 'county_me' and request.user.county:
        counties = County.objects.filter(id=request.user.county.id)
    
    context = {
        'counties': counties,
        'quarters': quarters,
        'thematic_areas': thematic_areas,
        'indicators': indicators,
    }
    return render(request, 'reports/generate.html', context)

@login_required
@view_reports_required
def quarterly_report(request):
    """Generate quarterly report - APPROVED ONLY"""
    user = request.user
    
    quarter_id = request.GET.get('quarter')
    county_id = request.GET.get('county')
    
    quarter = get_object_or_404(Quarter, id=quarter_id) if quarter_id else Quarter.objects.filter(is_active=True).first()
    
    if not quarter:
        return render(request, 'reports/quarterly.html', {'error': 'No active quarter found'})
    
    # Build query - APPROVED ONLY
    query = Q(quarter=quarter, status='approved')
    
    # ===== RBAC: COUNTY USERS ONLY SEE THEIR COUNTY =====
    if county_id:
        query &= Q(county_id=county_id)
        county = get_object_or_404(County, id=county_id)
    elif user.role and user.role.name == 'county_me' and user.county:
        query &= Q(county=user.county)
        county = user.county
    else:
        county = None
    
    entries = DataEntry.objects.filter(query).select_related('county', 'indicator', 'indicator__thematic_area')
    
    # Group by thematic area
    thematic_data = {}
    for entry in entries:
        area_name = entry.indicator.thematic_area.name
        if area_name not in thematic_data:
            thematic_data[area_name] = {
                'total': 0,
                'met': 0,
                'not_met': 0,
                'no_data': 0,
                'entries': []
            }
        thematic_data[area_name]['total'] += 1
        if entry.is_met() is True:
            thematic_data[area_name]['met'] += 1
        elif entry.is_met() is False:
            thematic_data[area_name]['not_met'] += 1
        else:
            thematic_data[area_name]['no_data'] += 1
        thematic_data[area_name]['entries'].append(entry)
    
    # County performance
    county_performance = {}
    for entry in entries:
        county_name = entry.county.name
        if county_name not in county_performance:
            county_performance[county_name] = {
                'total': 0,
                'met': 0,
                'not_met': 0,
                'entries': []
            }
        county_performance[county_name]['total'] += 1
        if entry.is_met() is True:
            county_performance[county_name]['met'] += 1
        elif entry.is_met() is False:
            county_performance[county_name]['not_met'] += 1
        county_performance[county_name]['entries'].append(entry)
    
    # Calculate overall
    total_entries = entries.count()
    total_met = sum(1 for e in entries if e.is_met() is True)
    total_not_met = sum(1 for e in entries if e.is_met() is False)
    total_no_data = sum(1 for e in entries if e.is_met() is None)
    
    context = {
        'quarter': quarter,
        'county': county,
        'thematic_data': thematic_data,
        'county_performance': county_performance,
        'total_entries': total_entries,
        'total_met': total_met,
        'total_not_met': total_not_met,
        'total_no_data': total_no_data,
        'user_role': user.role,
    }
    return render(request, 'reports/quarterly.html', context)

@login_required
@view_reports_required
def annual_report(request):
    """Generate annual report - APPROVED ONLY"""
    user = request.user
    year = request.GET.get('year', timezone.now().year)
    county_id = request.GET.get('county')
    
    quarters = Quarter.objects.filter(
        start_date__year=year
    ).order_by('start_date')
    
    if not quarters:
        return render(request, 'reports/annual.html', {
            'error': f'No quarters found for year {year}',
            'year': year
        })
    
    # Build query - APPROVED ONLY
    query = Q(quarter__in=quarters, status='approved')
    
    # ===== RBAC: COUNTY USERS ONLY SEE THEIR COUNTY =====
    if county_id:
        query &= Q(county_id=county_id)
        county = get_object_or_404(County, id=county_id)
    elif user.role and user.role.name == 'county_me' and user.county:
        query &= Q(county=user.county)
        county = user.county
    else:
        county = None
    
    entries = DataEntry.objects.filter(query).select_related('county', 'indicator', 'quarter')
    
    # Quarter-by-quarter performance
    quarterly_performance = {}
    for q in quarters:
        q_entries = entries.filter(quarter=q)
        quarterly_performance[q.name] = {
            'total': q_entries.count(),
            'met': sum(1 for e in q_entries if e.is_met() is True),
            'not_met': sum(1 for e in q_entries if e.is_met() is False),
            'no_data': sum(1 for e in q_entries if e.is_met() is None),
        }
    
    # Thematic performance for the year
    thematic_performance = {}
    for area in ThematicArea.objects.all():
        area_entries = entries.filter(indicator__thematic_area=area)
        met_count = sum(1 for e in area_entries if e.is_met() is True)
        thematic_performance[area.name] = {
            'total': area_entries.count(),
            'met': met_count,
            'not_met': sum(1 for e in area_entries if e.is_met() is False),
            'no_data': sum(1 for e in area_entries if e.is_met() is None),
            'percentage': round((met_count / area_entries.count() * 100) if area_entries.count() > 0 else 0)
        }
    
    total_entries = entries.count()
    total_met = sum(1 for e in entries if e.is_met() is True)
    total_not_met = sum(1 for e in entries if e.is_met() is False)
    total_no_data = sum(1 for e in entries if e.is_met() is None)
    
    context = {
        'year': year,
        'county': county,
        'quarters': quarters,
        'quarterly_performance': quarterly_performance,
        'thematic_performance': thematic_performance,
        'total_entries': total_entries,
        'total_met': total_met,
        'total_not_met': total_not_met,
        'total_no_data': total_no_data,
        'overall_percentage': round((total_met / total_entries * 100) if total_entries > 0 else 0),
        'user_role': user.role,
    }
    return render(request, 'reports/annual.html', context)

@login_required
@view_reports_required
def thematic_report(request, code):
    """Generate report for a specific thematic area - APPROVED ONLY"""
    user = request.user
    area = get_object_or_404(ThematicArea, code=code)
    indicators = Indicator.objects.filter(thematic_area=area, is_active=True)
    
    quarter_id = request.GET.get('quarter')
    county_id = request.GET.get('county')
    
    # Build query - APPROVED ONLY
    query = Q(indicator__in=indicators, status='approved')
    if quarter_id:
        query &= Q(quarter_id=quarter_id)
    if county_id:
        query &= Q(county_id=county_id)
    elif user.role and user.role.name == 'county_me' and user.county:
        query &= Q(county=user.county)
    
    entries = DataEntry.objects.filter(query).select_related('county', 'quarter', 'indicator')
    
    # Group by indicator
    indicator_performance = {}
    for ind in indicators:
        ind_entries = entries.filter(indicator=ind)
        met_count = sum(1 for e in ind_entries if e.is_met() is True)
        indicator_performance[ind.code] = {
            'name': ind.name,
            'target': ind.target_value,
            'unit': ind.unit,
            'total': ind_entries.count(),
            'met': met_count,
            'not_met': sum(1 for e in ind_entries if e.is_met() is False),
            'no_data': sum(1 for e in ind_entries if e.is_met() is None),
            'entries': ind_entries,
            'percentage': round((met_count / ind_entries.count() * 100) if ind_entries.count() > 0 else 0)
        }
    
    total_entries = entries.count()
    total_met = sum(1 for e in entries if e.is_met() is True)
    total_not_met = sum(1 for e in entries if e.is_met() is False)
    total_no_data = sum(1 for e in entries if e.is_met() is None)
    
    # Get available filters
    quarters = Quarter.objects.filter(is_active=True)
    counties = County.objects.filter(is_active=True)
    if user.role == 'county_me' and user.county:
        counties = County.objects.filter(id=user.county.id)
    
    context = {
        'area': area,
        'indicators': indicators,
        'indicator_performance': indicator_performance,
        'total_entries': total_entries,
        'total_met': total_met,
        'total_not_met': total_not_met,
        'total_no_data': total_no_data,
        'quarters': quarters,
        'counties': counties,
        'selected_quarter': quarter_id,
        'selected_county': county_id,
        'user_role': user.role,
    }
    return render(request, 'reports/thematic.html', context)

@login_required
@view_reports_required
def sdg_report(request):
    """SDG Indicators Report - APPROVED ONLY"""
    user = request.user
    
    # SDG-related indicators (based on your M&E framework)
    sdg_indicators = Indicator.objects.filter(
        Q(code__startswith='FERT-') | 
        Q(code__startswith='MM-') |
        Q(code__startswith='PHED-'),
        is_active=True
    )
    
    quarter_id = request.GET.get('quarter')
    county_id = request.GET.get('county')
    
    # Build query - APPROVED ONLY
    query = Q(indicator__in=sdg_indicators, status='approved')
    if quarter_id:
        query &= Q(quarter_id=quarter_id)
    if county_id:
        query &= Q(county_id=county_id)
    elif user.role and user.role.name == 'county_me' and user.county:
        query &= Q(county=user.county)
    
    entries = DataEntry.objects.filter(query).select_related('county', 'quarter', 'indicator')
    
    # Group by SDG goal (simplified mapping)
    sdg_mapping = {
        'FERT': {'goal': 'SDG 3 & 5', 'description': 'Good Health & Gender Equality'},
        'MM': {'goal': 'SDG 3', 'description': 'Good Health and Well-being'},
        'PHED': {'goal': 'SDG 11 & 13', 'description': 'Sustainable Cities & Climate Action'},
    }
    
    sdg_performance = {}
    for area_code, info in sdg_mapping.items():
        area_indicators = sdg_indicators.filter(thematic_area__code=area_code)
        area_entries = entries.filter(indicator__in=area_indicators)
        met_count = sum(1 for e in area_entries if e.is_met() is True)
        total = area_entries.count()
        not_met = sum(1 for e in area_entries if e.is_met() is False)
        no_data = sum(1 for e in area_entries if e.is_met() is None)
        
        # Handle division by zero
        percentage = round((met_count / total * 100) if total > 0 else 0)
        not_met_percentage = round((not_met / total * 100) if total > 0 else 0)
        
        sdg_performance[area_code] = {
            'goal': info['goal'],
            'description': info['description'],
            'indicators': area_indicators.count(),
            'total': total,
            'met': met_count,
            'not_met': not_met,
            'no_data': no_data,
            'percentage': percentage,
            'not_met_percentage': not_met_percentage,
        }
    
    total_entries = entries.count()
    total_met = sum(1 for e in entries if e.is_met() is True)
    
    # Get available filters
    quarters = Quarter.objects.filter(is_active=True)
    counties = County.objects.filter(is_active=True)
    if user.role and user.role.name == 'county_me' and user.county:
        counties = County.objects.filter(id=user.county.id)
    
    context = {
        'sdg_performance': sdg_performance,
        'total_entries': total_entries,
        'total_met': total_met,
        'overall_percentage': round((total_met / total_entries * 100) if total_entries > 0 else 0),
        'quarters': quarters,
        'counties': counties,
        'selected_quarter': quarter_id,
        'selected_county': county_id,
        'user_role': user.role.name if user.role else 'No Role',
    }
    return render(request, 'reports/sdg_report.html', context)

@login_required
@ncpd_or_admin_required
def pending_reports(request):
    """View pending approvals summary - For NCPD/Admin only"""
    user = request.user
    
    # Check if user is authorized for pending approvals
    if not user.is_superuser and (not user.role or user.role.name not in ['admin', 'ncpd_me']):
        return render(request, 'errors/403.html', status=403)
    
    pending_entries = DataEntry.objects.filter(status='submitted').select_related('county', 'quarter', 'indicator')
    
    # Group by county
    county_pending = {}
    for entry in pending_entries:
        county_name = entry.county.name
        if county_name not in county_pending:
            county_pending[county_name] = {
                'count': 0,
                'entries': []
            }
        county_pending[county_name]['count'] += 1
        county_pending[county_name]['entries'].append(entry)
    
    # Group by quarter
    quarter_pending = {}
    for entry in pending_entries:
        quarter_name = entry.quarter.name
        if quarter_name not in quarter_pending:
            quarter_pending[quarter_name] = {
                'count': 0,
                'entries': []
            }
        quarter_pending[quarter_name]['count'] += 1
        quarter_pending[quarter_name]['entries'].append(entry)
    
    context = {
        'pending_entries': pending_entries,
        'county_pending': county_pending,
        'quarter_pending': quarter_pending,
        'total_pending': pending_entries.count(),
    }
    return render(request, 'reports/pending.html', context)

@login_required
@view_reports_required
def export_report(request, format):
    """Export report in specified format - APPROVED ONLY"""
    user = request.user
    
    county_id = request.GET.get('county')
    quarter_id = request.GET.get('quarter')
    thematic_area_id = request.GET.get('thematic_area')
    
    query = Q(status='approved')
    if county_id:
        query &= Q(county_id=county_id)
    if quarter_id:
        query &= Q(quarter_id=quarter_id)
    if thematic_area_id:
        indicators = Indicator.objects.filter(thematic_area_id=thematic_area_id)
        query &= Q(indicator__in=indicators)
    
    # RBAC: County users only see their county
    if user.role == 'county_me' and user.county:
        query &= Q(county=user.county)
    
    entries = DataEntry.objects.filter(query).select_related('county', 'quarter', 'indicator')
    
    if format == 'csv':
        return export_csv(entries)
    elif format == 'json':
        return export_json(entries)
    else:
        return HttpResponse("Unsupported format", status=400)

def export_csv(entries):
    """Export entries as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(['County', 'Quarter', 'Indicator Code', 'Indicator Name', 'Value', 'Unit', 'Target', 'Status', 'Met Target'])
    
    for entry in entries:
        writer.writerow([
            entry.county.name,
            entry.quarter.name,
            entry.indicator.code,
            entry.indicator.name,
            entry.value or 'No Data',
            entry.indicator.unit,
            entry.indicator.target_value or 'N/A',
            entry.get_status_display(),
            'Yes' if entry.is_met() is True else 'No' if entry.is_met() is False else 'N/A'
        ])
    
    return response

def export_json(entries):
    """Export entries as JSON"""
    data = []
    for entry in entries:
        data.append({
            'county': entry.county.name,
            'quarter': entry.quarter.name,
            'indicator_code': entry.indicator.code,
            'indicator_name': entry.indicator.name,
            'value': entry.value or 'No Data',
            'unit': entry.indicator.unit,
            'target': str(entry.indicator.target_value) if entry.indicator.target_value else 'N/A',
            'status': entry.get_status_display(),
            'met_target': entry.is_met() is True,
            'submitted_at': entry.submitted_at.isoformat() if entry.submitted_at else None,
        })
    
    return JsonResponse(data, safe=False)

@login_required
@view_reports_required
def export_data(request):
    """Main export page with options"""
    user = request.user
    
    # Get data for filters
    counties = County.objects.filter(is_active=True)
    quarters = Quarter.objects.filter(is_active=True)
    thematic_areas = ThematicArea.objects.all()
    indicators = Indicator.objects.filter(is_active=True)
    
    # ===== RBAC: COUNTY USERS ONLY SEE THEIR COUNTY =====
    if user.role and user.role.name == 'county_me' and user.county:
        counties = County.objects.filter(id=user.county.id)
        # Only show indicators that have data for this county
        county_entries = DataEntry.objects.filter(county=user.county, status='approved')
        indicators = indicators.filter(
            id__in=county_entries.values_list('indicator_id', flat=True).distinct()
        )
        thematic_areas = thematic_areas.filter(
            id__in=indicators.values_list('thematic_area_id', flat=True).distinct()
        )
    
    context = {
        'counties': counties,
        'quarters': quarters,
        'thematic_areas': thematic_areas,
        'indicators': indicators,
    }
    return render(request, 'reports/export.html', context)

@login_required
def export_excel(request):
    """Export data to Excel with formatting"""
    user = request.user
    
    # Get filters from request
    county_id = request.GET.get('county')
    quarter_id = request.GET.get('quarter')
    thematic_area_id = request.GET.get('thematic_area')
    indicator_id = request.GET.get('indicator')
    export_type = request.GET.get('type', 'entries')
    
    # Build query
    query = Q()
    if county_id:
        query &= Q(county_id=county_id)
    if quarter_id:
        query &= Q(quarter_id=quarter_id)
    if thematic_area_id:
        indicators = Indicator.objects.filter(thematic_area_id=thematic_area_id)
        query &= Q(indicator__in=indicators)
    if indicator_id:
        query &= Q(indicator_id=indicator_id)
    
    # RBAC
    if user.role == 'county_me' and user.county:
        query &= Q(county=user.county)
    
    # Get entries
    entries = DataEntry.objects.filter(query).select_related('county', 'quarter', 'indicator')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Export"
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a5632", end_color="1a5632", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['County', 'Quarter', 'Indicator Code', 'Indicator Name', 'Thematic Area', 
               'Value', 'Unit', 'Target', 'Status', 'Met Target', 'Submitted By', 'Submitted At']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row, entry in enumerate(entries, 2):
        ws.cell(row=row, column=1, value=entry.county.name)
        ws.cell(row=row, column=2, value=entry.quarter.name)
        ws.cell(row=row, column=3, value=entry.indicator.code)
        ws.cell(row=row, column=4, value=entry.indicator.name)
        ws.cell(row=row, column=5, value=entry.indicator.thematic_area.name)
        ws.cell(row=row, column=6, value=entry.value or 'No Data')
        ws.cell(row=row, column=7, value=entry.indicator.unit)
        ws.cell(row=row, column=8, value=float(entry.indicator.target_value) if entry.indicator.target_value else 'N/A')
        ws.cell(row=row, column=9, value=entry.get_status_display())
        
        met_cell = ws.cell(row=row, column=10)
        if entry.is_met() is True:
            met_cell.value = 'Yes'
            met_cell.fill = PatternFill(start_color="e8f5e9", end_color="e8f5e9", fill_type="solid")
        elif entry.is_met() is False:
            met_cell.value = 'No'
            met_cell.fill = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
        else:
            met_cell.value = 'N/A'
        
        ws.cell(row=row, column=11, value=entry.submitted_by.username if entry.submitted_by else '')
        ws.cell(row=row, column=12, value=entry.submitted_at.strftime('%Y-%m-%d %H:%M') if entry.submitted_at else '')
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 18
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_indicators(request):
    """Export indicators to Excel"""
    indicators = Indicator.objects.filter(is_active=True).select_related('thematic_area')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicators"
    
    # Headers
    headers = ['Code', 'Name', 'Thematic Area', 'Type', 'Data Type', 'Unit', 'Target', 'Source', 'Frequency']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="1a5632", end_color="1a5632", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, ind in enumerate(indicators, 2):
        ws.cell(row=row, column=1, value=ind.code)
        ws.cell(row=row, column=2, value=ind.name)
        ws.cell(row=row, column=3, value=ind.thematic_area.name)
        ws.cell(row=row, column=4, value=ind.get_indicator_type_display())
        ws.cell(row=row, column=5, value=ind.get_data_type_display())
        ws.cell(row=row, column=6, value=ind.unit)
        ws.cell(row=row, column=7, value=float(ind.target_value) if ind.target_value else '')
        ws.cell(row=row, column=8, value=ind.source_system)
        ws.cell(row=row, column=9, value=ind.get_frequency_display())
    
    # Auto-adjust
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="indicators_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
@admin_required
def import_indicators(request):
    """Import indicators from Excel/CSV"""
    if request.method != 'POST':
        messages.error(request, 'Invalid request method.')
        return redirect('reports:export_data')
    
    if 'file' not in request.FILES:
        messages.error(request, 'No file uploaded.')
        return redirect('reports:export_data')
    
    file = request.FILES['file']
    filename = file.name.lower()
    
    try:
        import csv
        import io
        from openpyxl import load_workbook
        
        # Read the file
        if filename.endswith('.csv'):
            decoded = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(decoded))
            data = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            wb = load_workbook(file)
            ws = wb.active
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(row):
                    row_dict = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            row_dict[header] = row[i]
                    data.append(row_dict)
        else:
            messages.error(request, 'Unsupported file format. Please upload CSV or Excel.')
            return redirect('reports:export_data')
        
        # Process data
        imported = 0
        updated = 0
        errors = 0
        error_messages = []
        
        # Thematic area mapping
        thematic_map = {
            'FERT': 'Fertility',
            'MM': 'Morbidity & Mortality',
            'MIG': 'Migration & Urbanization',
            'PHED': 'Population, Health, Environment & Disaster'
        }
        
        for row in data:
            try:
                # Clean and get values - handle None values
                def get_value(key, default=''):
                    val = row.get(key)
                    if val is None:
                        return default
                    return str(val).strip()
                
                code = get_value('Code')
                name = get_value('Name')
                thematic_code = get_value('Thematic Area').upper()
                
                if not code or not name:
                    errors += 1
                    error_messages.append(f"Missing code or name")
                    continue
                
                # Get thematic area
                thematic_area = ThematicArea.objects.filter(code=thematic_code).first()
                if not thematic_area:
                    thematic_name = thematic_map.get(thematic_code, '')
                    thematic_area = ThematicArea.objects.filter(name__iexact=thematic_name).first()
                
                if not thematic_area:
                    errors += 1
                    error_messages.append(f"Thematic area not found: {thematic_code}")
                    continue
                
                # Helper to safely get numeric values
                def get_numeric(val):
                    if val is None or str(val).strip() == '':
                        return None
                    try:
                        return float(str(val).strip())
                    except ValueError:
                        return None
                
                # Helper to get string values
                def get_str(val):
                    if val is None:
                        return ''
                    return str(val).strip()
                
                # Check if exists
                if Indicator.objects.filter(code=code).exists():
                    indicator = Indicator.objects.get(code=code)
                    indicator.name = name
                    indicator.thematic_area = thematic_area
                    indicator.indicator_type = get_str(row.get('Type', 'output'))
                    indicator.data_type = get_str(row.get('Data Type', 'numeric'))
                    indicator.unit = get_str(row.get('Unit'))
                    indicator.target_value = get_numeric(row.get('Target'))
                    indicator.source_system = get_str(row.get('Source'))
                    indicator.frequency = get_str(row.get('Frequency', 'annual'))
                    indicator.min_value = get_numeric(row.get('Min Value'))
                    indicator.max_value = get_numeric(row.get('Max Value'))
                    indicator.description = get_str(row.get('Description'))
                    indicator.save()
                    updated += 1
                else:
                    Indicator.objects.create(
                        code=code,
                        name=name,
                        thematic_area=thematic_area,
                        indicator_type=get_str(row.get('Type', 'output')),
                        data_type=get_str(row.get('Data Type', 'numeric')),
                        unit=get_str(row.get('Unit')),
                        target_value=get_numeric(row.get('Target')),
                        source_system=get_str(row.get('Source')),
                        frequency=get_str(row.get('Frequency', 'annual')),
                        min_value=get_numeric(row.get('Min Value')),
                        max_value=get_numeric(row.get('Max Value')),
                        description=get_str(row.get('Description')),
                        created_by=request.user,
                    )
                    imported += 1
            except Exception as e:
                errors += 1
                error_messages.append(str(e))
        
        if imported > 0:
            messages.success(request, f'Imported {imported} new indicators.')
        if updated > 0:
            messages.success(request, f'Updated {updated} existing indicators.')
        if errors > 0:
            messages.error(request, f'{errors} errors occurred.')
            for msg in error_messages[:5]:
                messages.warning(request, msg)
        
        if imported == 0 and updated == 0 and errors == 0:
            messages.warning(request, 'No indicators were imported. Please check your file format.')
        
    except Exception as e:
        messages.error(request, f'Error processing file: {str(e)}')
    
    return redirect('reports:export_data')

@login_required
def download_template(request, template_type):
    """Download import template for different data types"""
    
    if template_type == 'indicators':
        return download_indicators_template()
    elif template_type == 'data_entries':
        return download_data_entries_template()
    elif template_type == 'partners':
        return download_partners_template()
    elif template_type == 'projects':
        return download_projects_template()
    else:
        messages.error(request, 'Invalid template type.')
        return redirect('reports:export_data')

def download_indicators_template():
    """Download template for importing indicators"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    from django.http import HttpResponse
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Indicators Template"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1a5632", end_color="1a5632", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers with descriptions
    headers = [
        ('Code*', 'Unique indicator code e.g., FERT-01'),
        ('Name*', 'Full indicator name'),
        ('Thematic Area*', 'FERT, MM, MIG, or PHED'),
        ('Type', 'impact, outcome, output, or action'),
        ('Data Type', 'numeric, percentage, decimal, boolean, or count'),
        ('Unit', 'e.g., %, per 100,000'),
        ('Target', 'Target value'),
        ('Source', 'Data source e.g., KNBS'),
        ('Frequency', 'annual, quarterly, monthly, 5_years, 10_years'),
        ('Min Value', 'Minimum allowed value'),
        ('Max Value', 'Maximum allowed value'),
        ('Description', 'Brief description')
    ]
    
    # Add headers
    for col, (header, comment) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Add sample data
    samples = [
        ['FERT-01', 'Human Development Index', 'FERT', 'impact', 'decimal', 'Index (0-1)', '0.70', 'UNDP', 'annual', '0', '1', 'Composite index of human development'],
        ['FERT-02', 'Population Growth Rate', 'FERT', 'outcome', 'percentage', '%', '2.0', 'KNBS', 'annual', '0', '5', 'Annual population growth rate'],
        ['FERT-03', 'Total Fertility Rate (National)', 'FERT', 'outcome', 'decimal', 'Births per woman', '3.0', 'KNBS', '5_years', '0', '8', 'Average number of children per woman'],
        ['MM-01', 'Life Expectancy at Birth', 'MM', 'outcome', 'numeric', 'Years', '70', 'KNBS', 'annual', '40', '85', 'Average life expectancy'],
        ['MM-02', 'Maternal Mortality Ratio', 'MM', 'output', 'numeric', 'Per 100,000', '70', 'MOH', 'annual', '0', '1000', 'Maternal deaths per 100,000 live births'],
        ['MIG-01', 'Migration Rate', 'MIG', 'outcome', 'percentage', '%', '', 'KNBS', 'annual', '0', '100', 'Rate of population migration'],
        ['PHED-01', 'Disaster & Climate Change Mortality Rate', 'PHED', 'outcome', 'numeric', 'Per 100,000', '', 'NDOC', 'annual', '0', '1000', 'Mortality rate related to disasters'],
    ]
    
    # Add sample data rows
    for row, data in enumerate(samples, 2):
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="indicators_import_template.xlsx"'
    wb.save(response)
    return response

def download_data_entries_template():
    """Download template for importing data entries"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data Entries Template"
    
    headers = [
        ('County*', 'County name exactly as in system'),
        ('Quarter*', 'Quarter name e.g., Q1 2025'),
        ('Indicator Code*', 'Indicator code e.g., FERT-01'),
        ('Value*', 'The actual data value'),
        ('Notes', 'Optional notes about this entry')
    ]
    
    for col, (header, comment) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1a5632", end_color="1a5632", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sample data
    samples = [
        ['Nairobi', 'Q1 2025', 'FERT-01', '0.72', 'From census data'],
        ['Mombasa', 'Q1 2025', 'FERT-02', '2.5', ''],
        ['Kisumu', 'Q1 2025', 'MM-01', '68', ''],
    ]
    
    for row, data in enumerate(samples, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="data_entries_import_template.xlsx"'
    wb.save(response)
    return response

def download_partners_template():
    """Download template for importing partners"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Partners Template"
    
    headers = [
        ('Code*', 'Unique partner code e.g., P-001'),
        ('Name*', 'Organization name'),
        ('Type', 'ngo, cbo, fbo, government, private, academic, development, other'),
        ('Contact Person*', 'Primary contact name'),
        ('Contact Email*', 'Email address'),
        ('Contact Phone', 'Phone number'),
        ('Address', 'Physical address'),
        ('Website', 'Website URL'),
        ('Status', 'active, inactive, pending, suspended'),
        ('Counties', 'Comma-separated county codes')
    ]
    
    for col, (header, comment) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1a5632", end_color="1a5632", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    samples = [
        ['P-001', 'Red Cross Kenya', 'ngo', 'John Doe', 'john@redcross.org', '+254-712-345-678', 'Nairobi', 'www.redcross.org', 'active', '001,047'],
        ['P-002', 'UNFPA Kenya', 'development', 'Jane Smith', 'jane@unfpa.org', '+254-712-345-679', 'Nairobi', 'www.unfpa.org', 'active', '001'],
    ]
    
    for row, data in enumerate(samples, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="partners_import_template.xlsx"'
    wb.save(response)
    return response

def download_projects_template():
    """Download template for importing projects"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects Template"
    
    headers = [
        ('Code*', 'Unique project code e.g., PRJ-001'),
        ('Name*', 'Project name'),
        ('Partner Code*', 'Partner code this project belongs to'),
        ('Start Date*', 'YYYY-MM-DD format'),
        ('End Date*', 'YYYY-MM-DD format'),
        ('Budget', 'Total budget in KSh'),
        ('Status', 'planning, active, completed, suspended, cancelled'),
        ('Project Lead', 'Name of project lead'),
        ('Project Email', 'Project email'),
        ('Project Phone', 'Project phone'),
        ('Indicators', 'Comma-separated indicator codes'),
        ('Counties', 'Comma-separated county codes')
    ]
    
    for col, (header, comment) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1a5632", end_color="1a5632", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    samples = [
        ['PRJ-001', 'Population Health Project', 'P-001', '2025-01-01', '2025-12-31', '5000000', 'active', 'John Lead', 'project@example.com', '+254-712-345-678', 'FERT-01,MM-01', '001,047'],
        ['PRJ-002', 'Migration Study', 'P-002', '2025-02-01', '2025-10-31', '3000000', 'planning', 'Jane Lead', 'migration@example.com', '+254-712-345-679', 'MIG-01', '002'],
    ]
    
    for row, data in enumerate(samples, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="projects_import_template.xlsx"'
    wb.save(response)
    return response